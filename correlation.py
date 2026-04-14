"""
Spearman correlation analysis between qPCR gene expression
and behavioural outcomes across therapy groups.

Groups:
    Control+CoCl₂, PTSD+CoCl₂  (stage_1 in behavioural files)
    Control+IHT,   PTSD+IHT     (stage_2)
    Control+Bar,   PTSD+Bar     (stage_3)

Correlations are computed on GROUP MEANS (n = 6 data points per pair).

Input files (same directory):
    PTSD_Hypoxia_Cortex.xlsx
    PTSD_Hypoxia_Hipo.xlsx
    /mnt/user-data/uploads/EP_DF_ALL.xlsx      — Elevated Plus-Maze
    /mnt/user-data/uploads/DL_DF_ALL.xlsx      — Dark-Light Box
    /mnt/user-data/uploads/OF_DF_ALL.xlsx      — Open Field

Output:
    correlation_input_data.csv  — all 170 tested pairs with raw group means
    correlations.png / .pdf     — scatter plots for significant pairs (p < 0.05)
"""

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy import stats
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ── Config ────────────────────────────────────────────────────────────────────
GROUP_OFFSETS = {
    'Control':       0,
    'PTSD':          6,
    'Control+CoCl₂': 12,
    'PTSD+CoCl₂':    18,
    'Control+IHT':   24,
    'PTSD+IHT':      30,
    'Control+Bar':   36,
    'PTSD+Bar':      42,
}
THERAPY_GROUPS = [
    'Control+CoCl₂', 'PTSD+CoCl₂',
    'Control+IHT',   'PTSD+IHT',
    'Control+Bar',   'PTSD+Bar',
]
SHEETS      = ['HIF1', 'HIF2', 'HIF3', 'PACAP', 'PAI1']
GENE_LABELS = dict(zip(SHEETS, ['HIF-1α', 'HIF-2α', 'HIF-3α', 'PACAP', 'PAI-1']))
DATA_ROW    = 6

PCR_FILES = {
    'Cortex':      'PTSD_Hypoxia_Cortex.xlsx',
    'Hippocampus': 'PTSD_Hypoxia_Hipo.xlsx',
}
BEH_FILES = {
    'EP': '/mnt/user-data/uploads/EP_DF_ALL.xlsx',
    'DL': '/mnt/user-data/uploads/DL_DF_ALL.xlsx',
    'OF': '/mnt/user-data/uploads/OF_DF_ALL.xlsx',
}
BEH_COLS = {
    'EP': ['distance (m)', 'freezing_episodes (n)', 'time_freezing (s)',
           'open_entries (n)', 'open_time (s)',
           'closed_entries (n)', 'closed_time (s)'],
    'DL': ['entries to light (n)', 'total_out (s)', 'total_in (s)', 'head_poking (n)'],
    'OF': ['distance (m)', 'mean_speed (m/s)', 'freezing_episodes (n)',
           'time_freezing (s)', 'center_entries (n)', 'center_time (s)'],
}

GROUP_COLORS = {
    'Control+CoCl₂': '#64B5F6', 'PTSD+CoCl₂':  '#1565C0',
    'Control+IHT':   '#81C784', 'PTSD+IHT':    '#2E7D32',
    'Control+Bar':   '#FFD54F', 'PTSD+Bar':    '#E65100',
}
GROUP_MARKERS = {
    'Control+CoCl₂': 'o', 'PTSD+CoCl₂':  's',
    'Control+IHT':   'o', 'PTSD+IHT':    's',
    'Control+Bar':   'o', 'PTSD+Bar':    's',
}
BEH_LABELS = {
    'closed_time (s)':     'Closed arm time (s)',
    'closed_entries (n)':  'Closed arm entries (n)',
    'mean_speed (m/s)':    'Mean speed (m/s)',
    'distance (m)':        'Distance (m)',
    'time_freezing (s)':   'Freezing time (s)',
    'head_poking (n)':     'Head-poking (n)',
}
TEST_LABELS = {
    'EP': 'Elevated Plus-Maze',
    'OF': 'Open Field',
    'DL': 'Dark-Light Box',
}

plt.rcParams.update({'font.family': 'DejaVu Sans', 'pdf.fonttype': 42, 'ps.fonttype': 42})


# ── PCR helpers ───────────────────────────────────────────────────────────────
def remove_iqr(vals, factor=1.5):
    if len(vals) < 4:
        return vals
    q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
    iqr = q3 - q1
    return [v for v in vals if q1 - factor * iqr <= v <= q3 + factor * iqr]


def load_pcr_vals(filepath, sheet, group):
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    ws   = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    off  = GROUP_OFFSETS[group]
    vals = []
    for r in rows[DATA_ROW:]:
        ac = r[off + 1] if len(r) > off + 1 else None
        gc = r[off + 2] if len(r) > off + 2 else None
        if (ac is not None and gc is not None
                and isinstance(ac, (int, float))
                and isinstance(gc, (int, float))):
            vals.append(2 ** (float(ac) - float(gc)) * 100)
    return remove_iqr(vals)


def pcr_group_means(tissue):
    fp     = PCR_FILES[tissue]
    means  = {}
    for sh in SHEETS:
        gene = GENE_LABELS[sh]
        for grp in THERAPY_GROUPS:
            vals = load_pcr_vals(fp, sh, grp)
            if vals:
                means[(gene, grp)] = np.mean(vals)
    return means


# ── Behavioural helpers ───────────────────────────────────────────────────────
def load_beh(test_key):
    fname = BEH_FILES[test_key]
    wb    = openpyxl.load_workbook(fname, read_only=True)
    ws    = list(wb.worksheets)[0]
    rows  = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr  = rows[0]
    data = [{h: v for h, v in zip(hdr, r) if h is not None}
            for r in rows[1:] if r[0] is not None]
    df   = pd.DataFrame(data)

    df['therapy'] = df['treatment'].map({
        'stage_1': 'CoCl₂', 'stage_2': 'IHT', 'stage_3': 'Bar', 'non': 'non'
    })
    df['group_clean'] = (df['group'].str.lower().str.strip()
                                    .str.replace('control', 'cont'))
    df = df[df['therapy'] != 'non'].copy()

    def _label(row):
        base = 'Control' if 'cont' in str(row['group_clean']) else 'PTSD'
        return f"{base}+{row['therapy']}"

    df['pcr_group'] = df.apply(_label, axis=1)

    skip = {'animal №', 'group', 'treatment', 'therapy', 'group_clean', 'pcr_group'}
    for col in df.columns:
        if col not in skip:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def beh_group_means(test_key):
    df    = load_beh(test_key)
    means = {}
    for col in BEH_COLS[test_key]:
        if col not in df.columns:
            continue
        for grp, val in df.groupby('pcr_group')[col].mean().items():
            means[(col, grp)] = val
    return means


# ── Correlation engine ────────────────────────────────────────────────────────
def run_correlations():
    results = []

    for tissue in ['Cortex', 'Hippocampus']:
        pcr = pcr_group_means(tissue)

        for test_key in ['EP', 'DL', 'OF']:
            beh = beh_group_means(test_key)

            for sh in SHEETS:
                gene = GENE_LABELS[sh]

                for col in BEH_COLS[test_key]:
                    common = [g for g in THERAPY_GROUPS
                              if (gene, g) in pcr and (col, g) in beh]
                    if len(common) < 5:
                        continue

                    x = [pcr[(gene, g)]  for g in common]
                    y = [beh[(col,  g)]  for g in common]

                    if np.std(x) < 1e-8 or np.std(y) < 1e-8:
                        continue

                    r, p = stats.spearmanr(x, y)

                    row = {
                        'tissue':      tissue,
                        'gene':        gene,
                        'beh_test':    test_key,
                        'beh_variable': col,
                        'n_groups':    len(common),
                        'spearman_r':  round(r, 4),
                        'p_value':     round(p, 4),
                        'significant': 'yes' if p < 0.05 else 'no',
                    }
                    for g, xv, yv in zip(common, x, y):
                        gk = (g.replace('+', '_')
                               .replace('₂', '2').replace('₃', '3'))
                        row[f'PCR_{gk}'] = round(xv, 4)
                        row[f'BEH_{gk}'] = round(yv, 4)
                    results.append(row)

    return pd.DataFrame(results).sort_values('p_value').reset_index(drop=True)


# ── Plot ──────────────────────────────────────────────────────────────────────
def plot_correlations(df_results, out_prefix='correlations'):
    sig = df_results[df_results['significant'] == 'yes'].copy()
    if sig.empty:
        print('No significant correlations to plot.')
        return

    pcr_cache = {t: pcr_group_means(t) for t in ['Cortex', 'Hippocampus']}
    beh_cache = {k: beh_group_means(k) for k in ['EP', 'DL', 'OF']}

    n     = len(sig)
    ncols = 4
    nrows = -(-n // ncols)

    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(ncols * 4.2, nrows * 4.0),
                             squeeze=False)
    fig.patch.set_facecolor('white')

    for idx, row in sig.iterrows():
        ax      = axes[idx // ncols][idx % ncols]
        tissue  = row['tissue']
        gene    = row['gene']
        test    = row['beh_test']
        col     = row['beh_variable']

        pcr = pcr_cache[tissue]
        beh = beh_cache[test]

        common = [g for g in THERAPY_GROUPS
                  if (gene, g) in pcr and (col, g) in beh]
        x = [pcr[(gene, g)] for g in common]
        y = [beh[(col,  g)] for g in common]

        r, p = stats.spearmanr(x, y)

        m, b = np.polyfit(x, y, 1)
        xline = np.linspace(min(x), max(x), 100)
        ax.plot(xline, m * xline + b, '--', color='#555555',
                linewidth=1.2, alpha=0.7, zorder=1)

        for g, xi, yi in zip(common, x, y):
            ax.scatter(xi, yi,
                       color=GROUP_COLORS[g], marker=GROUP_MARKERS[g],
                       s=110, edgecolors='#333333', linewidths=0.8,
                       zorder=3, label=g.replace('Control', 'Ctrl'))

        p_str = f'p = {p:.4f}' if p >= 0.001 else 'p < 0.001'
        ax.set_title(f'{tissue}\n{gene}  ↔  {TEST_LABELS[test]}',
                     fontsize=9, fontweight='bold', color='#212121', pad=6)
        ax.set_xlabel(f'{gene} [C]rel', fontsize=8.5, color='#424242')
        ax.set_ylabel(BEH_LABELS.get(col, col), fontsize=8.5, color='#424242')
        ax.text(0.96, 0.96, f'ρ = {r:.3f}\n{p_str}',
                transform=ax.transAxes, ha='right', va='top', fontsize=8.5,
                bbox=dict(boxstyle='round,pad=0.3', fc='white',
                          ec='#CCCCCC', alpha=0.85))

        ax.set_facecolor('#F5F5F5')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#BDBDBD')
        ax.spines['bottom'].set_color('#BDBDBD')
        ax.tick_params(labelsize=8)
        ax.yaxis.grid(True, color='white', linewidth=0.8, zorder=0)
        ax.xaxis.grid(True, color='white', linewidth=0.8, zorder=0)
        ax.set_axisbelow(True)

    for i in range(n, nrows * ncols):
        axes[i // ncols][i % ncols].set_visible(False)

    from matplotlib.lines import Line2D
    handles = [
        Line2D([0], [0], marker=GROUP_MARKERS[g], color='w',
               markerfacecolor=GROUP_COLORS[g], markeredgecolor='#333333',
               markersize=9, label=g.replace('Control', 'Ctrl'))
        for g in THERAPY_GROUPS
    ]
    fig.legend(handles=handles, loc='lower right', ncol=2, fontsize=8,
               framealpha=0.9, bbox_to_anchor=(0.99, 0.01))

    fig.suptitle(
        'Spearman correlations between gene expression ([C]rel) '
        'and behavioural outcomes\n'
        'Group means (n = 6 groups). Dashed line: linear trend. '
        'Circles = Control, Squares = PTSD background.',
        fontsize=10, color='#212121', y=1.01, fontweight='bold'
    )

    plt.tight_layout(rect=[0, 0, 1, 1])
    plt.subplots_adjust(wspace=0.42, hspace=0.62)
    fig.savefig(f'{out_prefix}.png', dpi=300, bbox_inches='tight', facecolor='white')
    fig.savefig(f'{out_prefix}.pdf', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'Figure → {out_prefix}.png / .pdf')


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Running Spearman correlations...')
    df = run_correlations()

    df.to_csv('correlation_input_data.csv', index=False)
    print(f'CSV → correlation_input_data.csv  ({len(df)} pairs tested)')

    sig = df[df['significant'] == 'yes']
    print(f'\nSignificant pairs (p < 0.05): {len(sig)}')
    print(sig[['tissue', 'gene', 'beh_test', 'beh_variable',
               'spearman_r', 'p_value']].to_string(index=False))

    plot_correlations(df, out_prefix='correlations')
    print('\nDone.')