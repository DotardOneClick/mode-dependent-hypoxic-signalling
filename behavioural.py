"""
Behavioural Statistical Analysis — Article 2
Boxplot + jitter, ggplot2 style. Separate figures for OF, EPM, DLB.
Control (non) and PTSD (non) shown as reference median lines.
6 therapy groups shown as boxplots.
Statistical pipeline:
    1. Outlier removal — IQR 1.5×
    2. Normality — Shapiro-Wilk per group
    3. Overall — one-way ANOVA (all normal) or Kruskal-Wallis
    4. Post-hoc — t-test (after ANOVA) or Mann-Whitney U (after Kruskal)
Usage:
    python behavioural_art2.py
    python behavioural_art2.py --test EPM
    python behavioural_art2.py --no-outliers
Data files in data/ folder:
    data/OF_DF_ALL.xlsx
    data/EP_DF_ALL.xlsx
    data/DL_DF_ALL.xlsx
"""
import os, csv, argparse
from datetime import datetime
from collections import defaultdict
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from scipy import stats
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# ── Config ──────────────────────────────────────────────────────────────────────
DATA_FILES = {
    'OF':  'data/OF_DF_ALL.xlsx',
    'EPM': 'data/EP_DF_ALL.xlsx',
    'DLB': 'data/DL_DF_ALL.xlsx',
}
EPM_MAP = {'cont': 'control'}
REF_KEYS = {
    'OF':  {'Control': ('control', 'non'), 'PTSD': ('ptsd', 'non')},
    'EPM': {'Control': ('control', 'non'), 'PTSD': ('ptsd', 'non')},
    'DLB': {'Control': ('control', 'non'), 'PTSD': ('ptsd', 'non')},
}
THERAPY_KEYS = {
    'Control+CoCl₂': ('control', 'stage_1'),
    'PTSD+CoCl₂':    ('ptsd',    'stage_1'),
    'Control+IHT':   ('control', 'stage_2'),
    'PTSD+IHT':      ('ptsd',    'stage_2'),
    'Control+Bar':   ('control', 'stage_3'),
    'PTSD+Bar':      ('ptsd',    'stage_3'),
}
THERAPY_COLORS = {
    'Control+CoCl₂': '#64B5F6',
    'PTSD+CoCl₂':    '#1565C0',
    'Control+IHT':   '#81C784',
    'PTSD+IHT':      '#2E7D32',
    'Control+Bar':   '#FFD54F',
    'PTSD+Bar':      '#E65100',
}
CTRL_LINE_COLOR = '#388E3C'
PTSD_LINE_COLOR = '#7B1FA2'
BOX_POSITIONS = [0, 1, 3, 4, 6, 7]
BOX_GROUPS    = list(THERAPY_KEYS.keys())
PARAMS = {
    'OF': {
        'distance (m)':          'Distance (m)',
        'mean_speed (m/s)':      'Speed (m/s)',
        'freezing_episodes (n)': 'Freezing episodes (n)',
        'time_freezing (s)':     'Freezing time (s)',
        'center_entries (n)':    'Center entries (n)',
        'center_time (s)':       'Center time (s)',
        'corners_entries (n)':   'Corner entries (n)',
        'corners_time (s)':      'Corner time (s)',
        'sides_entries (n)':     'Side entries (n)',
        'sides_time (s)':        'Side time (s)',
    },
    'EPM': {
        'distance (m)':          'Distance (m)',
        'time_freezing (s)':     'Freezing time (s)',
        'freezing_episodes (n)': 'Freezing episodes (n)',
        'open_entries (n)':      'Open arm entries (n)',
        'open_time (s)':         'Open arm time (s)',
        'open_head_entries (n)': 'Open head dips (n)',
        'closed_entries (n)':    'Closed arm entries (n)',
        'closed_time (s)':       'Closed arm time (s)',
        'rotations (n)':         'Rotations (n)',
    },
    'DLB': {
        'entries to light (n)': 'Light zone entries (n)',
        'total_out (s)':        'Time in light (s)',
        'total_in (s)':         'Time in dark (s)',
        'head_poking (n)':      'Head poking (n)',
    },
}
COMPARISON_PAIRS = [
    ('Control', 'Control+CoCl₂', 'Ctrl+CoCl₂ vs Control'),
    ('Control', 'PTSD+CoCl₂',    'PTSD+CoCl₂ vs Control'),
    ('PTSD',    'PTSD+CoCl₂',    'PTSD+CoCl₂ vs PTSD'),
    ('Control+CoCl₂', 'PTSD+CoCl₂', 'PTSD+CoCl₂ vs Ctrl+CoCl₂'),
    ('Control', 'Control+IHT',   'Ctrl+IHT vs Control'),
    ('Control', 'PTSD+IHT',      'PTSD+IHT vs Control'),
    ('PTSD',    'PTSD+IHT',      'PTSD+IHT vs PTSD'),
    ('Control+IHT', 'PTSD+IHT',  'PTSD+IHT vs Ctrl+IHT'),
    ('Control', 'Control+Bar',   'Ctrl+Bar vs Control'),
    ('Control', 'PTSD+Bar',      'PTSD+Bar vs Control'),
    ('PTSD',    'PTSD+Bar',      'PTSD+Bar vs PTSD'),
    ('Control+Bar', 'PTSD+Bar',  'PTSD+Bar vs Ctrl+Bar'),
]
plt.rcParams.update({
    'font.family':  'DejaVu Sans',
    'pdf.fonttype': 42,
    'ps.fonttype':  42,
})

# ── Data loading ────────────────────────────────────────────────────────────────
def load_data(filepath, test_name):
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    ws   = wb.active
    cols = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        g = str(row[1]).lower()
        t = str(row[2]).lower()
        if g in EPM_MAP:
            g = EPM_MAP[g]
        data[(g, t)].append(row)
    return cols, data

def get_vals(data, cols, key_tuple, param):
    if param not in cols: return []
    ci = cols.index(param)
    return [float(r[ci]) for r in data.get(key_tuple, [])
            if r[ci] is not None and isinstance(r[ci], (int, float))]

def remove_iqr(vals, factor=1.5):
    if len(vals) < 4: return vals
    q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
    iqr = q3 - q1
    return [v for v in vals if q1 - factor*iqr <= v <= q3 + factor*iqr]

# ── Statistics ──────────────────────────────────────────────────────────────────
def shapiro_wilk(v):
    if len(v) < 3: return False, None
    _, p = stats.shapiro(v)
    return p >= 0.05, round(p, 4)

def run_overall(groups_dict):
    vals = list(groups_dict.values())
    norm = all(shapiro_wilk(v)[0] for v in vals if len(v) >= 3)
    if norm:
        stat, p = stats.f_oneway(*vals)
        return {'test': 'One-way ANOVA', 'statistic': round(stat, 4),
                'p': round(p, 6), 'all_normal': True}
    stat, p = stats.kruskal(*vals)
    return {'test': 'Kruskal-Wallis', 'statistic': round(stat, 4),
            'p': round(p, 6), 'all_normal': False}

def run_posthoc(v1, v2, use_ttest):
    if len(v1) < 3 or len(v2) < 3: return None, 'n/a'
    if use_ttest:
        _, p = stats.ttest_ind(v1, v2)
        return round(p, 6), 't-test'
    _, p = stats.mannwhitneyu(v1, v2, alternative='two-sided')
    return round(p, 6), 'Mann-Whitney U'

def sig_label(p):
    if p is None: return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'

# ── Plot ─────────────────────────────────────────────────────────────────────────
def plot_param(ax, ref_data, box_data, label, posthoc, show_ylabel=False):
    x_min, x_max = -0.7, 7.7

    # ── Reference lines ────────────────────────────────────────────────────────
    ctrl_vals   = ref_data.get('Control', [])
    ptsd_vals   = ref_data.get('PTSD', [])
    ctrl_median = np.median(ctrl_vals) if ctrl_vals else None
    ptsd_median = np.median(ptsd_vals) if ptsd_vals else None

    if ctrl_median is not None:
        ax.hlines(ctrl_median, x_min, x_max,
                  colors=CTRL_LINE_COLOR, linestyles='--',
                  linewidth=2.0, zorder=6)
    if ptsd_median is not None:
        ax.hlines(ptsd_median, x_min, x_max,
                  colors=PTSD_LINE_COLOR, linestyles=':',
                  linewidth=2.0, zorder=6)

    # ── Boxplots ───────────────────────────────────────────────────────────────
    all_vals = list(ctrl_vals) + list(ptsd_vals)
    for pos, grp in zip(BOX_POSITIONS, BOX_GROUPS):
        vals = box_data.get(grp, [])
        if not vals: continue
        all_vals += vals
        color = THERAPY_COLORS[grp]

        bp = ax.boxplot(
            [vals], positions=[pos], widths=0.55,
            patch_artist=True, notch=False,
            medianprops=dict(color='#212121', linewidth=2.0),
            whiskerprops=dict(color='#424242', linewidth=1.2),
            capprops=dict(color='#424242', linewidth=1.2),
            flierprops=dict(marker='', markersize=0),
            boxprops=dict(linewidth=1.2),
            zorder=3,
        )
        bp['boxes'][0].set_facecolor(color)
        bp['boxes'][0].set_alpha(0.75)
        bp['boxes'][0].set_edgecolor('#212121')

        np.random.seed(42 + pos)
        jitter = np.random.uniform(-0.15, 0.15, len(vals))
        ax.scatter(pos + jitter, vals, color='#212121',
                   s=22, alpha=0.65, linewidths=0, zorder=5)

    # ── Y limits ───────────────────────────────────────────────────────────────
    if not all_vals: return
    data_max = max(all_vals)
    data_min = min(all_vals)
    top    = data_max * 1.60
    bottom = -top * 0.06 if data_min >= 0 else data_min * 1.1
    ax.set_ylim(bottom, top)
    ax.set_xlim(x_min, x_max)

    # ── Significance brackets ──────────────────────────────────────────────────
    sig_pairs = [(g1, g2, r) for (g1, g2), r in posthoc.items()
                 if r.get('sig', '') not in ('ns', '')]
    sig_pairs.sort(key=lambda item: _span(item[0], item[1]))

    level_tops = {}
    base_y = data_max * 1.08
    step   = (top - base_y) / max(len(sig_pairs) + 2, 5)

    for g1, g2, result in sig_pairs:
        sl = result.get('sig', '')
        p1 = _get_x(g1)
        p2 = _get_x(g2)
        if p1 is None or p2 is None: continue
        lo, hi = min(p1, p2), max(p1, p2)

        y = base_y
        for col in [lo, hi]:
            if col in level_tops:
                y = max(y, level_tops[col] + step)
        for col in [lo, hi]:
            level_tops[col] = y + step * 0.6

        needed = y + step * 1.4
        if needed > ax.get_ylim()[1]:
            ax.set_ylim(bottom, needed)

        h = step * 0.22
        ax.plot([p1, p1, p2, p2], [y, y+h, y+h, y],
                lw=1.0, color='#212121', clip_on=False, zorder=8)
        ax.text((p1+p2)/2, y + h*1.3, sl,
                ha='center', va='bottom', fontsize=9,
                color='#212121', fontweight='bold',
                clip_on=False, zorder=9)

    # ── Formatting ────────────────────────────────────────────────────────────
    ax.set_xticks(BOX_POSITIONS)
    ax.set_xticklabels(
        [g.replace('+', '\n+') for g in BOX_GROUPS],
        fontsize=8, fontweight='bold', color='#212121', linespacing=1.2
    )
    ax.set_title(label, fontsize=11, fontweight='bold', color='#212121', pad=8)
    if show_ylabel:
        ax.set_ylabel('Value', fontsize=11, fontweight='bold',
                      color='#424242', labelpad=12)

    ax.set_facecolor('#EBEBEB')
    ax.yaxis.grid(True, color='white', linewidth=1.0, zorder=0)
    ax.xaxis.grid(False)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#BDBDBD')
    ax.spines['bottom'].set_linewidth(0.8)
    ax.tick_params(axis='x', length=0, pad=10)
    ax.tick_params(axis='y', length=0, labelsize=9)


def _get_x(group_name):
    if group_name in BOX_GROUPS:
        return BOX_POSITIONS[BOX_GROUPS.index(group_name)]
    if group_name in ('Control', 'PTSD'):
        return -0.5
    return None


def _span(g1, g2):
    p1 = _get_x(g1) if _get_x(g1) is not None else -1
    p2 = _get_x(g2) if _get_x(g2) is not None else -1
    return abs(p1 - p2)


def make_figure(test_name, ref_data_all, box_data_all,
                posthoc_all, output_prefix):
    params  = PARAMS[test_name]
    n       = len(params)
    ncols   = 5
    nrows   = -(-n // ncols)
    figsize = (ncols * 5.0, nrows * 7.0)

    fig, axes = plt.subplots(nrows, ncols, figsize=figsize)
    fig.patch.set_facecolor('white')
    axes_flat = np.array(axes).flatten() if n > 1 else [axes]

    for i, (param, label) in enumerate(params.items()):
        ax = axes_flat[i]
        if param not in box_data_all:
            ax.set_visible(False)
            continue
        plot_param(ax,
                   ref_data_all.get(param, {}),
                   box_data_all[param],
                   label,
                   posthoc_all.get(param, {}),
                   show_ylabel=(i % ncols == 0))

    for idx in range(n, len(axes_flat)):
        axes_flat[idx].set_visible(False)

    legend_elements = [
        mlines.Line2D([0], [0], color=CTRL_LINE_COLOR, linestyle='--',
                      linewidth=2, label='Control (median)'),
        mlines.Line2D([0], [0], color=PTSD_LINE_COLOR, linestyle=':',
                      linewidth=2, label='PTSD (median)'),
    ] + [
        mpatches.Patch(facecolor=THERAPY_COLORS[g], edgecolor='#212121',
                       alpha=0.75, label=g) for g in BOX_GROUPS
    ]
    fig.legend(handles=legend_elements, loc='lower center', ncol=4,
               fontsize=9, frameon=True, fancybox=False, edgecolor='#E0E0E0',
               bbox_to_anchor=(0.5, -0.03))

    titles = {'OF': 'Open Field Test', 'EPM': 'Elevated Plus Maze',
              'DLB': 'Dark-Light Box'}
    fig.suptitle(f'{titles[test_name]} — Article 2',
                 fontsize=14, fontweight='bold', color='#212121', y=1.02)
    fig.text(0.5, -0.06,
             'Dashed/dotted lines = Control/PTSD baseline medians. '
             'Boxplot: median, IQR, whiskers = 1.5×IQR. '
             'Dots = individual animals (IQR outliers removed). '
             '* p<0.05, ** p<0.01, *** p<0.001.',
             ha='center', fontsize=8, color='#9E9E9E', style='italic')

    plt.tight_layout(rect=[0, 0.06, 1, 1])
    plt.subplots_adjust(wspace=0.40, hspace=0.65)
    fig.savefig(f'{output_prefix}.png', dpi=300,
                bbox_inches='tight', facecolor='white')
    fig.savefig(f'{output_prefix}.pdf', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f"  Figure → {output_prefix}.png / .pdf")

# ── CSV ─────────────────────────────────────────────────────────────────────────
def save_csv(rows, filepath):
    if not rows: return
    fields = list(rows[0].keys())
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"  CSV    → {filepath}")

# ── Main ─────────────────────────────────────────────────────────────────────────
def run_test(test_name, out_dir, apply_iqr=True):
    filepath = DATA_FILES[test_name]
    params   = PARAMS[test_name]

    print(f"\n{'='*60}")
    print(f"  {test_name}  |  IQR: {'ON' if apply_iqr else 'OFF'}")
    print(f"{'='*60}")

    cols, data = load_data(filepath, test_name)

    ref_data_all = {}
    box_data_all = {}
    posthoc_all  = {}
    csv_rows     = []

    for param, label in params.items():
        ref = {}
        for ref_name, key in REF_KEYS[test_name].items():
            v = get_vals(data, cols, key, param)
            ref[ref_name] = remove_iqr(v) if apply_iqr else v

        box = {}
        for grp_name, key in THERAPY_KEYS.items():
            v = get_vals(data, cols, key, param)
            box[grp_name] = remove_iqr(v) if apply_iqr else v

        if not any(len(v) > 0 for v in box.values()):
            continue

        ref_data_all[param] = ref
        box_data_all[param] = box

        all_groups = {**ref, **box}
        non_empty  = {g: v for g, v in all_groups.items() if len(v) >= 3}
        if len(non_empty) < 2: continue
        overall = run_overall(non_empty)

        posthoc = {}
        for g1, g2, pair_label in COMPARISON_PAIRS:
            v1 = all_groups.get(g1, [])
            v2 = all_groups.get(g2, [])
            if len(v1) < 3 or len(v2) < 3: continue
            p, test = run_posthoc(v1, v2, overall['all_normal'])
            sl = sig_label(p)
            posthoc[(g1, g2)] = {
                'p': p, 'sig': sl, 'test': test, 'label': pair_label
            }
            if sl not in ('ns', ''):
                d = '↑' if np.mean(v2) > np.mean(v1) else '↓'
                print(f"  {label:<25} {pair_label:<35} {sl} p={p:.4f} {d}")

        posthoc_all[param] = posthoc

        for g1, g2, pair_label in COMPARISON_PAIRS:
            v1 = all_groups.get(g1, [])
            v2 = all_groups.get(g2, [])
            if not v1 or not v2: continue
            ph = posthoc.get((g1, g2), {})
            m1, m2 = np.mean(v1), np.mean(v2)
            s1, s2 = np.std(v1, ddof=1), np.std(v2, ddof=1)
            csv_rows.append({
                'test':         test_name,
                'parameter':    label,
                'comparison':   pair_label,
                'group1':       g1,
                'group2':       g2,
                'overall_test': overall['test'],
                'overall_p':    overall['p'],
                'posthoc_test': ph.get('test', ''),
                'posthoc_p':    ph.get('p', ''),
                'significance': ph.get('sig', ''),
                'direction':    ('↑' if m2>m1 else '↓') if ph.get('sig','') not in ('ns','') else '—',
                'mean_g1':      round(m1, 4),
                'sd_g1':        round(s1, 4),
                'sem_g1':       round(s1/np.sqrt(len(v1)), 4),
                'n_g1':         len(v1),
                'mean_g2':      round(m2, 4),
                'sd_g2':        round(s2, 4),
                'sem_g2':       round(s2/np.sqrt(len(v2)), 4),
                'n_g2':         len(v2),
            })

    prefix = os.path.join(out_dir, f'Art2_{test_name}')
    make_figure(test_name, ref_data_all, box_data_all, posthoc_all, prefix)
    save_csv(csv_rows, os.path.join(out_dir, f'Art2_{test_name}_stats.csv'))


def main():
    parser = argparse.ArgumentParser(
        description='Behavioural analysis — Article 2')
    parser.add_argument('--test', choices=['OF', 'EPM', 'DLB', 'ALL'],
                        default='ALL')
    parser.add_argument('--no-outliers', action='store_true')
    args = parser.parse_args()

    ts      = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    out_dir = os.path.join('results', ts)
    os.makedirs(out_dir, exist_ok=True)
    print(f"Output → {out_dir}/")

    tests = ['OF', 'EPM', 'DLB'] if args.test == 'ALL' else [args.test]
    for t in tests:
        run_test(t, out_dir, apply_iqr=not args.no_outliers)

    print(f"\nDone! → {out_dir}/")


if __name__ == '__main__':
    main()