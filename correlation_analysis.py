"""
Spearman Group Correlations
Gene expression (qPCR, recalculated ΔCt→ΔΔCt→[C]rel pipeline) vs Behavioural parameters
Group means (n=7 groups). Molecular data: Grubbs-cleaned per-animal [C]rel, averaged per group.
Behavioural data: IQR 1.5× outlier removal on group values (unchanged).

Usage:
    python correlation.py

Data files in data/ folder:
    data/PCR_Results_Corrected.xlsx   (raw Ct, used via preprocess.py)
    data/OF_DF_ALL.xlsx
    data/EP_DF_ALL.xlsx
    data/DL_DF_ALL.xlsx
"""

import os, csv
from datetime import datetime
from collections import defaultdict

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
import openpyxl
import warnings
warnings.filterwarnings('ignore')

from preprocess import run as run_preprocess

# ── Config ──────────────────────────────────────────────────────────────────────

FILES_BEHAV = {
    'EPM': 'data/EP_DF_ALL.xlsx',
    'OF':  'data/OF_DF_ALL.xlsx',
    'DLB': 'data/DL_DF_ALL.xlsx',
}

GENES = ['HIF-1', 'HIF-2', 'HIF-3', 'PACAP', 'PAI']
GENE_LABELS = {
    'HIF-1': 'HIF-1α', 'HIF-2': 'HIF-2α', 'HIF-3': 'HIF-3α',
    'PACAP': 'PACAP',  'PAI': 'PAI-1',
}
TISSUES = ['mPFC', 'Hippocampus']

PLOT_GROUPS = [
    'Control', 'Control+CoCl₂', 'PTSD+CoCl₂',
    'Control+IHT', 'PTSD+IHT',
    'Control+Bar', 'PTSD+Bar',
]

GROUP_COLORS = {
    'Control':       '#9E9E9E',
    'Control+CoCl₂': '#64B5F6',
    'PTSD+CoCl₂':    '#1565C0',
    'Control+IHT':   '#81C784',
    'PTSD+IHT':      '#2E7D32',
    'Control+Bar':   '#FFD54F',
    'PTSD+Bar':      '#E65100',
}

EPM_MAP = {'cont': 'control', 'ptsd': 'ptsd'}

GROUP_KEYS = {
    'Control':        ('control', 'non'),
    'Control+CoCl₂':  ('control', 'stage_1'),
    'PTSD+CoCl₂':     ('ptsd',    'stage_1'),
    'Control+IHT':    ('control', 'stage_2'),
    'PTSD+IHT':       ('ptsd',    'stage_2'),
    'Control+Bar':    ('control', 'stage_3'),
    'PTSD+Bar':       ('ptsd',    'stage_3'),
}

# Behavioural parameters to correlate
BEHAV_PARAMS = {
    'EPM': {
        'time_freezing (s)':     'Freezing time EPM (s)',
        'open_time (s)':         'Open arm time EPM (s)',
        'open_head_entries (n)': 'Open head dips EPM (n)',
        'freezing_episodes (n)': 'Freezing episodes EPM (n)',
    },
    'OF': {
        'time_freezing (s)':  'Freezing time OF (s)',
        'corners_time (s)':   'Corner time OF (s)',
        'center_time (s)':    'Center time OF (s)',
    },
    'DLB': {
        'entries to light (n)': 'Light zone entries DLB (n)',
        'total_out (s)':        'Time in light DLB (s)',
        'total_in (s)':         'Time in dark DLB (s)',
    },
}

plt.rcParams.update({
    'font.family':  'DejaVu Sans',
    'pdf.fonttype': 42,
    'ps.fonttype':  42,
})

# ── Data loading ────────────────────────────────────────────────────────────────

def remove_iqr(vals, factor=1.5):
    """Used for behavioural group values only — molecular data is already
    Grubbs-cleaned at the per-animal ΔCt stage in preprocess.py."""
    if len(vals) < 4: return vals
    q1, q3 = np.percentile(vals, 25), np.percentile(vals, 75)
    iqr = q3 - q1
    return [v for v in vals if q1 - factor*iqr <= v <= q3 + factor*iqr]


def load_qpcr_group_means():
    """Returns {tissue: {gene: {group: mean}}} from the recalculated
    per-animal [C]rel pipeline (ΔCt -> Grubbs -> ΔΔCt -> [C]rel)."""
    crel_df, _ = run_preprocess()
    means = {t: {} for t in TISSUES}
    for tissue in TISSUES:
        for gene in GENES:
            label = GENE_LABELS[gene]
            sub = crel_df[(crel_df.tissue_label == tissue) & (crel_df.gene == gene)]
            means[tissue][label] = {
                g: (sub.loc[sub.group_label == g, 'Crel'].mean() if g in sub.group_label.values else None)
                for g in PLOT_GROUPS
            }
    return means


def load_behav_group_means(filepath, test_name, params):
    """Returns {param: {group: mean}} for all params and groups."""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    cols = [c.value for c in ws[1]]
    data = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        g = str(row[1]).lower()
        t = str(row[2]).lower()
        if test_name in ('EPM', 'DLB') and g in EPM_MAP:
            g = EPM_MAP[g]
        data[(g, t)].append(row)

    means = {}
    for param in params:
        if param not in cols: continue
        ci = cols.index(param)
        means[param] = {}
        for grp_name, (g, t) in GROUP_KEYS.items():
            vals = [float(r[ci]) for r in data.get((g, t), [])
                    if r[ci] is not None and isinstance(r[ci], (int, float))]
            vals = remove_iqr(vals)
            means[param][grp_name] = np.mean(vals) if vals else None
    return means

# ── Statistics ──────────────────────────────────────────────────────────────────

def sig_label(p):
    if p is None: return ''
    if p < 0.001: return '***'
    if p < 0.01:  return '**'
    if p < 0.05:  return '*'
    return 'ns'

# ── Plot ─────────────────────────────────────────────────────────────────────────

def plot_correlation(ax, xs, ys, grps, gene, tissue, blabel, r, p):
    sl = sig_label(p)

    # Scatter points colored by group
    for x, y, grp in zip(xs, ys, grps):
        ax.scatter(x, y, color=GROUP_COLORS[grp], s=130, zorder=5,
                   edgecolors='#212121', linewidths=0.8)

    # Regression line
    if len(xs) >= 2:
        m, b = np.polyfit(xs, ys, 1)
        x_line = np.linspace(min(xs) * 0.85, max(xs) * 1.15, 100)
        ax.plot(x_line, m * x_line + b,
                color='#424242', lw=1.5, linestyle='--', zorder=4)

    # r / p annotation
    ax.text(0.97, 0.97,
            f'r = {r:.3f}\np = {p:.4f} {sl}',
            transform=ax.transAxes, ha='right', va='top',
            fontsize=12, color='#212121',
            bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                      edgecolor='#E0E0E0', alpha=0.9))

    ax.set_xlabel(f'{gene} [C]rel ({tissue})',
                  fontsize=13, fontweight='bold', color='#212121', labelpad=6)
    ax.set_ylabel(blabel, fontsize=13, fontweight='bold',
                  color='#212121', labelpad=6)
    ax.set_title(f'{gene} ({tissue}) vs {blabel}',
                 fontsize=13, fontweight='bold', color='#212121', pad=8)

    ax.set_facecolor('#EBEBEB')
    ax.yaxis.grid(True, color='white', linewidth=1.0, zorder=0)
    ax.xaxis.grid(True, color='white', linewidth=1.0, zorder=0)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_color('#BDBDBD')
    ax.spines['bottom'].set_linewidth(0.8)
    ax.tick_params(axis='both', length=0, labelsize=11)


def make_correlation_figure(sig_results, output_prefix):
    """One figure with all significant correlations."""
    n = len(sig_results)
    if n == 0:
        print("  No significant correlations to plot.")
        return

    ncols = min(n, 3)
    nrows = -(-n // ncols)
    figsize = (ncols * 5.0, nrows * 4.6)

    fig, axes = plt.subplots(nrows, ncols, figsize=figsize)
    fig.patch.set_facecolor('white')
    axes_flat = np.array(axes).flatten() if n > 1 else [axes]

    for i, res in enumerate(sig_results):
        plot_correlation(
            axes_flat[i],
            res['xs'], res['ys'], res['groups'],
            res['gene'], res['tissue'], res['blabel'],
            res['r'], res['p']
        )

    for idx in range(n, len(axes_flat)):
        axes_flat[idx].set_visible(False)

    # Legend
    patches = [mpatches.Patch(facecolor=GROUP_COLORS[g], edgecolor='#212121',
                               label=g) for g in PLOT_GROUPS]
    fig.legend(handles=patches, loc='lower center', ncol=4,
               fontsize=12, frameon=True, fancybox=False, edgecolor='#E0E0E0',
               bbox_to_anchor=(0.5, -0.05))

    fig.suptitle(
        'Spearman correlations — Gene expression vs Behaviour',
        fontsize=16, fontweight='bold', color='#212121', y=1.01
    )
    fig.text(
        0.5, -0.09,
        'Each point = group mean [C]rel (n=7 groups). Dashed line = linear fit. '
        'Spearman rank correlation. * p<0.05, ** p<0.01, *** p<0.001.',
        ha='center', fontsize=10, color='#9E9E9E', style='italic'
    )

    plt.tight_layout(rect=[0.01, 0.02, 0.99, 0.98])
    plt.subplots_adjust(wspace=0.28, hspace=0.42)
    fig.savefig(f'{output_prefix}.png', dpi=300,
                bbox_inches='tight', facecolor='white', pad_inches=0.1)
    fig.savefig(f'{output_prefix}.pdf',
                bbox_inches='tight', facecolor='white', pad_inches=0.1)
    plt.close(fig)
    print(f"  Figure → {output_prefix}.png / .pdf")

# ── CSV ─────────────────────────────────────────────────────────────────────────

def save_csv(rows, filepath):
    if not rows: return
    fields = list(rows[0].keys())
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"  CSV    → {filepath}")

# ── Main ─────────────────────────────────────────────────────────────────────────

def run():
    ts      = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    out_dir = os.path.join('results', ts)
    os.makedirs(out_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Spearman Correlations")
    print(f"Output → {out_dir}/")
    print('='*60)

    # ── Load qPCR group means (recalculated pipeline) ──────────────────────────
    print("\nLoading qPCR data (ΔCt -> Grubbs -> ΔΔCt -> [C]rel)...")
    qpcr_means = load_qpcr_group_means()

    # ── Load behavioural group means ───────────────────────────────────────────
    print("Loading behavioural data...")
    behav_means = {}   # {param_key: {group: mean}}
    behav_labels = {}  # {param_key: label}
    for test_name, filepath in FILES_BEHAV.items():
        params = BEHAV_PARAMS[test_name]
        loaded = load_behav_group_means(filepath, test_name, params)
        for pk, label in params.items():
            if pk in loaded:
                behav_means[pk] = loaded[pk]
                behav_labels[pk] = label

    # ── Run all correlations ───────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print(f"  Running Spearman correlations (n=7 groups)...")
    print(f"{'─'*60}")
    print(f"\n  {'Tissue':<12} {'Gene':<10} {'Parameter':<30} {'r':>7} {'p':>8} {'Sig':>4}")
    print("  " + "-"*68)

    all_results  = []
    sig_results  = []
    csv_rows     = []

    for tissue in TISSUES:
        for gene_key in GENES:
            gene = GENE_LABELS[gene_key]
            gene_vals = [qpcr_means[tissue][gene].get(g) for g in PLOT_GROUPS]

            for bparam, blabel in behav_labels.items():
                behav_vals = [behav_means[bparam].get(g) for g in PLOT_GROUPS]

                # Filter out None pairs
                pairs = [(gv, bv, grp)
                         for gv, bv, grp in zip(gene_vals, behav_vals, PLOT_GROUPS)
                         if gv is not None and bv is not None]

                if len(pairs) < 5: continue
                xs, ys, grps = zip(*pairs)

                r, p = stats.spearmanr(xs, ys)
                sl   = sig_label(p)

                csv_rows.append({
                    'tissue':    tissue,
                    'gene':      gene,
                    'parameter': blabel,
                    'n_groups':  len(pairs),
                    'r':         round(r, 4),
                    'p':         round(p, 4),
                    'sig':       sl,
                })

                if sl != 'ns':
                    marker = ' ←'
                    print(f"  {tissue:<12} {gene:<10} {blabel:<30} "
                          f"{r:>7.3f} {p:>8.4f} {sl:>4}{marker}")
                    sig_results.append({
                        'tissue': tissue, 'gene': gene,
                        'bparam': bparam, 'blabel': blabel,
                        'xs': list(xs), 'ys': list(ys), 'groups': list(grps),
                        'r': r, 'p': p, 'sig': sl,
                    })

    print(f"\n  Significant: {len(sig_results)} / {len(csv_rows)} tested")

    # ── Save outputs ───────────────────────────────────────────────────────────
    prefix = os.path.join(out_dir, 'Art2_correlations')
    make_correlation_figure(sig_results, prefix)
    save_csv(csv_rows, os.path.join(out_dir, 'Art2_correlations_stats.csv'))

    print(f"\nDone! → {out_dir}/")


if __name__ == '__main__':
    run()